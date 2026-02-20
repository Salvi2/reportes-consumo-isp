"""
Sistema de Reportes de Consumo ISP
Módulo de Reporte Anual
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

from constantes import (
    MESES, TOTAL_CONTRATADO, NOMBRE_ARCHIVO,
    UMBRAL_VERDE, UMBRAL_AMARILLO,
    COLOR_VERDE, COLOR_AMARILLO, COLOR_ROJO,
    COLOR_BLANCO, COLOR_NEGRO,
    COLOR_PRINCIPAL_HEX, COLOR_SECUNDARIO_HEX,
    COLOR_ENCABEZADO, COLOR_FILA_TOTAL
)


def generar_reporte_anual(anio, datos_anio):

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Reporte Anual"

    # Estilos
    fuente_encabezado = Font(
        name="Segoe UI", size=11, bold=True,
        color=COLOR_BLANCO
    )
    fuente_datos = Font(name="Segoe UI", size=11)
    fuente_total = Font(name="Segoe UI", size=11, bold=True)
    fuente_sin_datos = Font(
        name="Segoe UI", size=11,
        color="AAAAAA"
    )

    alineacion_centro = Alignment(horizontal="center", vertical="center")
    alineacion_izquierda = Alignment(horizontal="left", vertical="center")

    relleno_encabezado = PatternFill(
        start_color=COLOR_SECUNDARIO_HEX,
        end_color=COLOR_SECUNDARIO_HEX,
        fill_type="solid"
    )
    relleno_total = PatternFill(
        start_color=COLOR_FILA_TOTAL,
        end_color=COLOR_FILA_TOTAL,
        fill_type="solid"
    )

    borde = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )

    # ═══════════════════════════════════════
    # ENCABEZADOS (fila 1)
    # ═══════════════════════════════════════
    encabezados = ["Mes", "Consumo (Mbps)", "Contratado (Mbps)", "Uso (%)"]

    for col in range(1, 5):
        celda = hoja.cell(row=1, column=col)
        celda.value = encabezados[col - 1]
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = alineacion_centro
        celda.border = borde

    # ═══════════════════════════════════════
    # DATOS POR MES (filas 2 a 13)
    # ═══════════════════════════════════════
    fila = 2
    suma_consumo = 0
    meses_con_datos = 0

    for mes in MESES:

        # Columna A: Mes
        celda_mes = hoja.cell(row=fila, column=1)
        celda_mes.value = mes
        celda_mes.alignment = alineacion_izquierda
        celda_mes.border = borde

        # Columna B: Consumo
        celda_consumo = hoja.cell(row=fila, column=2)
        celda_consumo.alignment = alineacion_centro
        celda_consumo.border = borde
        celda_consumo.number_format = "#,##0.00"

        # Columna C: Contratado
        celda_contratado = hoja.cell(row=fila, column=3)
        celda_contratado.alignment = alineacion_centro
        celda_contratado.border = borde
        celda_contratado.number_format = "#,##0"

        # Columna D: Porcentaje
        celda_porcentaje = hoja.cell(row=fila, column=4)
        celda_porcentaje.alignment = alineacion_centro
        celda_porcentaje.border = borde
        celda_porcentaje.number_format = "0.00"

        if mes in datos_anio:
            # Mes CON datos
            datos_mes = datos_anio[mes]

            celda_mes.font = fuente_datos
            celda_consumo.font = fuente_datos
            celda_contratado.font = fuente_datos

            celda_consumo.value = datos_mes["total_consumo"]
            celda_contratado.value = TOTAL_CONTRATADO
            celda_porcentaje.value = datos_mes["porcentaje_general"]

            # Formato condicional
            aplicar_color_celda(celda_porcentaje, datos_mes["porcentaje_general"])

            suma_consumo += datos_mes["total_consumo"]
            meses_con_datos += 1
        else:
            # Mes SIN datos
            celda_mes.font = fuente_sin_datos
            celda_consumo.font = fuente_sin_datos
            celda_contratado.font = fuente_sin_datos
            celda_porcentaje.font = fuente_sin_datos

            celda_consumo.value = "—"
            celda_contratado.value = "—"
            celda_porcentaje.value = "—"

        fila += 1

    # ═══════════════════════════════════════
    # FILA PROMEDIO (fila 14)
    # ═══════════════════════════════════════
    fila_total = 14

    celda_ta = hoja.cell(row=fila_total, column=1)
    celda_ta.value = "PROMEDIO ANUAL"
    celda_ta.font = fuente_total
    celda_ta.fill = relleno_total
    celda_ta.alignment = alineacion_izquierda
    celda_ta.border = borde

    celda_tb = hoja.cell(row=fila_total, column=2)
    celda_tb.font = fuente_total
    celda_tb.fill = relleno_total
    celda_tb.alignment = alineacion_centro
    celda_tb.border = borde
    celda_tb.number_format = "#,##0.00"

    celda_tc = hoja.cell(row=fila_total, column=3)
    celda_tc.value = TOTAL_CONTRATADO
    celda_tc.font = fuente_total
    celda_tc.fill = relleno_total
    celda_tc.alignment = alineacion_centro
    celda_tc.border = borde
    celda_tc.number_format = "#,##0"

    celda_td = hoja.cell(row=fila_total, column=4)
    celda_td.font = fuente_total
    celda_td.fill = relleno_total
    celda_td.alignment = alineacion_centro
    celda_td.border = borde
    celda_td.number_format = "0.00"

    if meses_con_datos > 0:
        promedio_consumo = round(suma_consumo / meses_con_datos, 2)
        promedio_porcentaje = round((promedio_consumo / TOTAL_CONTRATADO) * 100, 2)
        celda_tb.value = promedio_consumo
        celda_td.value = promedio_porcentaje
        aplicar_color_celda(celda_td, promedio_porcentaje)
    else:
        celda_tb.value = 0
        celda_td.value = 0

    # ═══════════════════════════════════════
    # GRÁFICO ANUAL
    # ═══════════════════════════════════════
    generar_grafico_anual(hoja, anio)

    # Ancho de columnas
    hoja.column_dimensions["A"].width = 18
    hoja.column_dimensions["B"].width = 20
    hoja.column_dimensions["C"].width = 22
    hoja.column_dimensions["D"].width = 12

    # Guardar
    nombre_archivo = f"{NOMBRE_ARCHIVO}_Anual_{anio}.xlsx"
    libro.save(nombre_archivo)

    return nombre_archivo


def aplicar_color_celda(celda, valor):
    if valor < UMBRAL_VERDE:
        celda.fill = PatternFill(
            start_color=COLOR_VERDE,
            end_color=COLOR_VERDE,
            fill_type="solid"
        )
        celda.font = Font(
            name="Segoe UI", size=11,
            bold=True, color=COLOR_NEGRO
        )
    elif valor <= UMBRAL_AMARILLO:
        celda.fill = PatternFill(
            start_color=COLOR_AMARILLO,
            end_color=COLOR_AMARILLO,
            fill_type="solid"
        )
        celda.font = Font(
            name="Segoe UI", size=11,
            bold=True, color=COLOR_NEGRO
        )
    else:
        celda.fill = PatternFill(
            start_color=COLOR_ROJO,
            end_color=COLOR_ROJO,
            fill_type="solid"
        )
        celda.font = Font(
            name="Segoe UI", size=11,
            bold=True, color=COLOR_BLANCO
        )


def generar_grafico_anual(hoja, anio):

    grafico = BarChart()
    grafico.type = "col"
    grafico.grouping = "clustered"
    grafico.title = f"Consumo vs Contratado - {anio}"
    grafico.y_axis.title = "Mbps"
    grafico.x_axis.title = "Mes"
    grafico.width = 22
    grafico.height = 14

    # Consumo (columna B)
    datos_consumo = Reference(hoja, min_col=2, min_row=1, max_row=13)

    # Contratado (columna C)
    datos_contratado = Reference(hoja, min_col=3, min_row=1, max_row=13)

    # Meses (columna A)
    categorias = Reference(hoja, min_col=1, min_row=2, max_row=13)

    grafico.add_data(datos_consumo, titles_from_data=True)
    grafico.add_data(datos_contratado, titles_from_data=True)
    grafico.set_categories(categorias)

    # Colores UNIMET
    grafico.series[0].graphicalProperties.solidFill = COLOR_PRINCIPAL_HEX   # Naranja
    grafico.series[1].graphicalProperties.solidFill = COLOR_SECUNDARIO_HEX  # Azul

    grafico.x_axis.delete = False
    grafico.y_axis.delete = False

    hoja.add_chart(grafico, "A16")