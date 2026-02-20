"""
Sistema de Reportes de Consumo ISP
Módulo de Generación de Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

from constantes import (
    PROVEEDORES, CONTRATADO, TOTAL_CONTRATADO,
    UMBRAL_VERDE, UMBRAL_AMARILLO,
    COLOR_VERDE, COLOR_AMARILLO, COLOR_ROJO,
    COLOR_BLANCO, COLOR_NEGRO,
    COLOR_ENCABEZADO, COLOR_FILA_TOTAL,
    COLOR_PRINCIPAL_HEX, COLOR_SECUNDARIO_HEX
)


def generar_reporte(mes, anio, consumos, porcentajes, total_consumo, porcentaje_general):

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Reporte"

    # ═══════════════════════════════════════
    # ESTILOS
    # ═══════════════════════════════════════
    fuente_encabezado = Font(
        name="Segoe UI", size=11, bold=True,
        color=COLOR_BLANCO
    )
    fuente_datos = Font(name="Segoe UI", size=11)
    fuente_total = Font(name="Segoe UI", size=11, bold=True)

    alineacion_centro = Alignment(horizontal="center", vertical="center")
    alineacion_izquierda = Alignment(horizontal="left", vertical="center")

    relleno_encabezado = PatternFill(
        start_color=COLOR_SECUNDARIO_HEX,
        end_color=COLOR_SECUNDARIO_HEX,
        fill_type="solid"
    )

    relleno_total = PatternFill(
        start_color=COLOR_FILA_TOTAL,
        end_color=COLOR_FILA_TOTAL,
        fill_type="solid"
    )

    borde = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )

    # ═══════════════════════════════════════
    # ENCABEZADOS (fila 1)
    # ═══════════════════════════════════════
    encabezados = ["Proveedor", "Contratado (Mbps)", "Consumo (Mbps)", "Uso (%)"]

    for col in range(1, 5):
        celda = hoja.cell(row=1, column=col)
        celda.value = encabezados[col - 1]
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = alineacion_centro
        celda.border = borde

    # ═══════════════════════════════════════
    # DATOS (filas 2 a 6)
    # ═══════════════════════════════════════
    for i in range(len(PROVEEDORES)):
        fila = 2 + i

        celda_a = hoja.cell(row=fila, column=1)
        celda_a.value = PROVEEDORES[i]
        celda_a.font = fuente_datos
        celda_a.alignment = alineacion_izquierda
        celda_a.border = borde

        celda_b = hoja.cell(row=fila, column=2)
        celda_b.value = CONTRATADO[i]
        celda_b.font = fuente_datos
        celda_b.alignment = alineacion_centro
        celda_b.border = borde
        celda_b.number_format = "#,##0"

        celda_c = hoja.cell(row=fila, column=3)
        celda_c.value = consumos[i]
        celda_c.font = fuente_datos
        celda_c.alignment = alineacion_centro
        celda_c.border = borde
        celda_c.number_format = "#,##0.00"

        celda_d = hoja.cell(row=fila, column=4)
        celda_d.value = porcentajes[i]
        celda_d.font = fuente_datos
        celda_d.alignment = alineacion_centro
        celda_d.border = borde
        celda_d.number_format = "0.00"

    # ═══════════════════════════════════════
    # FILA TOTAL (fila 7)
    # ═══════════════════════════════════════
    fila_total = 7

    celda_ta = hoja.cell(row=fila_total, column=1)
    celda_ta.value = "TOTAL"
    celda_ta.font = fuente_total
    celda_ta.fill = relleno_total
    celda_ta.alignment = alineacion_izquierda
    celda_ta.border = borde

    celda_tb = hoja.cell(row=fila_total, column=2)
    celda_tb.value = TOTAL_CONTRATADO
    celda_tb.font = fuente_total
    celda_tb.fill = relleno_total
    celda_tb.alignment = alineacion_centro
    celda_tb.border = borde
    celda_tb.number_format = "#,##0"

    celda_tc = hoja.cell(row=fila_total, column=3)
    celda_tc.value = total_consumo
    celda_tc.font = fuente_total
    celda_tc.fill = relleno_total
    celda_tc.alignment = alineacion_centro
    celda_tc.border = borde
    celda_tc.number_format = "#,##0.00"

    celda_td = hoja.cell(row=fila_total, column=4)
    celda_td.value = porcentaje_general
    celda_td.font = fuente_total
    celda_td.fill = relleno_total
    celda_td.alignment = alineacion_centro
    celda_td.border = borde
    celda_td.number_format = "0.00"

    # ═══════════════════════════════════════
    # FORMATO CONDICIONAL
    # ═══════════════════════════════════════
    aplicar_formato_condicional(hoja, porcentajes, porcentaje_general)

    # ═══════════════════════════════════════
    # GRÁFICO
    # ═══════════════════════════════════════
    generar_grafico(hoja, mes, anio)

    # ═══════════════════════════════════════
    # ANCHO DE COLUMNAS
    # ═══════════════════════════════════════
    hoja.column_dimensions["A"].width = 18
    hoja.column_dimensions["B"].width = 20
    hoja.column_dimensions["C"].width = 18
    hoja.column_dimensions["D"].width = 12

    # ═══════════════════════════════════════
    # GUARDAR
    # ═══════════════════════════════════════
    nombre_archivo = f"Reporte_ISP_{mes}_{anio}.xlsx"
    libro.save(nombre_archivo)

    return nombre_archivo


def aplicar_formato_condicional(hoja, porcentajes, porcentaje_general):

    todos = porcentajes + [porcentaje_general]

    for i in range(len(todos)):
        fila = 2 + i
        celda = hoja.cell(row=fila, column=4)
        valor = todos[i]

        if valor < UMBRAL_VERDE:
            celda.fill = PatternFill(
                start_color=COLOR_VERDE,
                end_color=COLOR_VERDE,
                fill_type="solid"
            )
            celda.font = Font(
                name="Segoe UI", size=11,
                bold=True, color=COLOR_NEGRO
            )

        elif valor <= UMBRAL_AMARILLO:
            celda.fill = PatternFill(
                start_color=COLOR_AMARILLO,
                end_color=COLOR_AMARILLO,
                fill_type="solid"
            )
            celda.font = Font(
                name="Segoe UI", size=11,
                bold=True, color=COLOR_NEGRO
            )

        else:
            celda.fill = PatternFill(
                start_color=COLOR_ROJO,
                end_color=COLOR_ROJO,
                fill_type="solid"
            )
            celda.font = Font(
                name="Segoe UI", size=11,
                bold=True, color=COLOR_BLANCO
            )


def generar_grafico(hoja, mes, anio):

    grafico = BarChart()
    grafico.type = "col"
    grafico.grouping = "clustered"
    grafico.title = f"Consumo vs Contratado - {mes} {anio}"
    grafico.y_axis.title = "Mbps"
    grafico.x_axis.title = "Proveedores"
    grafico.width = 20
    grafico.height = 12

    # Contratado (columna B, fila 1 = encabezado, filas 2-6 = datos)
    datos_contratado = Reference(
        hoja, min_col=2, min_row=1, max_row=6
    )

    # Consumido (columna C)
    datos_consumo = Reference(
        hoja, min_col=3, min_row=1, max_row=6
    )

    # AQUÍ ESTÁ EL PROBLEMA ↓
    # Los nombres de proveedores están en columna A, filas 2 a 6
    categorias = Reference(
        hoja, min_col=1, min_row=2, max_row=6
    )

    grafico.add_data(datos_contratado, titles_from_data=True)
    grafico.add_data(datos_consumo, titles_from_data=True)

    # Esta línea debe ir DESPUÉS de add_data
    grafico.set_categories(categorias)

    # Colores
    grafico.series[0].graphicalProperties.solidFill = COLOR_SECUNDARIO_HEX
    grafico.series[1].graphicalProperties.solidFill = COLOR_PRINCIPAL_HEX

    grafico.x_axis.delete = False    # Eje X: nombres
    grafico.y_axis.delete = False    # Eje Y: números

    hoja.add_chart(grafico, "A9")